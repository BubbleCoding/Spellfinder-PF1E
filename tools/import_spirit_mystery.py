"""Import oracle mystery bonus spells and shaman spirit magic spells.

Reads 'spirit and mystery.xlsx' from the project root.

Excel structure:
  Row  2:      Header "Shaman Spirit"
  Rows 3-18:   Shaman spirits  (col A = spirit name, col B = spell list text)
  Row  21:     Header "Oracle Mysteries"
  Rows 22-55:  Oracle mysteries (col A = mystery name, col B = spell list text)

Spell list formats:
  Spirits:    "Spirit Magic Spells: spell1 (1st), spell2 (2nd), ..."
  Mysteries:  "Bonus Spells: spell1 (2nd), spell2 (4th), ..."

Result:
  Updates the `spirit` and `mystery` TEXT columns on the spells table.
  Each column holds a comma-separated list of names, e.g. "Flame, Battle".
"""

import os
import re
import sqlite3
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    raise SystemExit("ERROR: openpyxl is required. Run: pip install openpyxl")

ROOT_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH   = os.path.join(ROOT_DIR, "pfinder.db")
XLSX_PATH = os.path.join(ROOT_DIR, "data", "spirit and mystery.xlsx")

# Matches a spell entry: any text followed by (Nth) where N is an integer
ENTRY_RE = re.compile(r'(.+?)\s*\((\d+)(?:st|nd|rd|th)\)', re.IGNORECASE)

# Known name mismatches between the AON spreadsheet and the DB
NAME_FIXES = {
    "repel metal and stone": "repel metal or stone",  # typo in source
    "wail of the banshees":  "wail of the banshee",   # typo in source
    "horrid withering":      "horrid wilting",         # typo in source
}


def clean_spell_name(name):
    """Strip source-book abbreviations glued to the end of spell names.

    e.g. "stone fistAPG"  -> "stone fist"
         "force punchUM"  -> "force punch"
    """
    return re.sub(r'([a-z])([A-Z]{2,5})$', r'\1', name.strip()).strip()


def parse_spell_list(text):
    """Extract [(raw_spell_name, level_int), ...] from a spell list string."""
    text = re.sub(r'^[^:]+:\s*', '', text).strip()
    results = []
    for m in ENTRY_RE.finditer(text):
        raw_name = m.group(1).strip().strip(',.').strip()
        level = int(m.group(2))
        results.append((raw_name, level))
    return results


def find_spell_id(cur, raw_name):
    """Look up a spell by name with several fallback strategies."""
    name = clean_spell_name(raw_name)

    # Normalise Unicode curly apostrophes to straight apostrophe
    name = name.replace('\u2019', "'").replace('\u2018', "'")

    # Apply known source-data name fixes (typos etc.)
    name = NAME_FIXES.get(name.lower(), name)

    def lookup(n):
        row = cur.execute(
            "SELECT id FROM spells WHERE LOWER(name) = ?", (n.lower(),)
        ).fetchone()
        return row[0] if row else None

    # 1. Exact case-insensitive match
    result = lookup(name)
    if result:
        return result

    # 2. Strip trailing parenthetical qualifier (e.g. "(fire elementals only)")
    base = re.sub(r'\s*\([^)]*\)\s*$', '', name).strip()
    if base != name:
        result = lookup(base)
        if result:
            return result

    # 3. Reverse "adjective noun" -> "noun, adjective" (PF1e naming convention)
    m = re.match(r'^(greater|lesser|mass)\s+(.+)$', name, re.IGNORECASE)
    if m:
        reordered = f"{m.group(2)}, {m.group(1)}"
        result = lookup(reordered)
        if result:
            return result

    return None


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"Excel file not found: {XLSX_PATH}")
        raise SystemExit(1)
    if not os.path.exists(DB_PATH):
        print(f"Database not found: {DB_PATH}")
        raise SystemExit(1)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    not_found = []
    spirit_map  = defaultdict(list)   # spell_id -> [spirit_name, ...]
    mystery_map = defaultdict(list)   # spell_id -> [mystery_name, ...]

    def process_rows(row_range, entry_type, spell_map):
        for r in row_range:
            col_a = ws[r][0].value
            col_b = ws[r][1].value
            if not col_a or not col_b:
                continue

            entry_name = str(col_a).strip()
            text       = str(col_b).strip()

            if entry_type == "spirit":
                if entry_name.lower().endswith(" spirit"):
                    entry_name = entry_name[:-7].strip()

            for raw_name, _ in parse_spell_list(text):
                spell_id = find_spell_id(cur, raw_name)
                if spell_id is None:
                    not_found.append((entry_type, entry_name, raw_name))
                    continue
                spell_map[spell_id].append(entry_name)

    process_rows(range(3, 19),  "spirit",  spirit_map)
    process_rows(range(22, 56), "mystery", mystery_map)

    # Reset existing spirit/mystery columns, then populate from collected data
    cur.execute("UPDATE spells SET spirit = NULL")
    cur.execute("UPDATE spells SET mystery = NULL")

    for spell_id, names in spirit_map.items():
        cur.execute(
            "UPDATE spells SET spirit = ? WHERE id = ?",
            (", ".join(sorted(set(names))), spell_id),
        )

    for spell_id, names in mystery_map.items():
        cur.execute(
            "UPDATE spells SET mystery = ? WHERE id = ?",
            (", ".join(sorted(set(names))), spell_id),
        )

    conn.commit()
    conn.close()

    spirit_assoc  = sum(len(v) for v in spirit_map.values())
    mystery_assoc = sum(len(v) for v in mystery_map.values())
    print(f"Updated {len(spirit_map)} spells with spirit tags "
          f"({spirit_assoc} spirit-spell associations across 16 spirits).")
    print(f"Updated {len(mystery_map)} spells with mystery tags "
          f"({mystery_assoc} mystery-spell associations across 34 mysteries).")

    if not_found:
        print(f"\nSpells not found in database ({len(not_found)}):")
        for entry_type, entry_name, spell_name in not_found:
            print(f"  [{entry_type}: {entry_name}] {spell_name!r}")


if __name__ == "__main__":
    main()
